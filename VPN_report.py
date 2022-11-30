import openpyxl
import re
import os
import csv
import zipfile
import time
from datetime import datetime
from multiprocessing.connection import wait
from sys import builtin_module_names
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import PyPDF2 


o = webdriver.ChromeOptions()
o.add_argument(r"--user-data-dir=C:\\Users\\mukulnag\\AppData\\Local\\Google\\Chrome\\User Data\\Default") #e.g. C:\Users\You\AppData\Local\Google\Chrome\User Data
o.add_argument(r'--profile-directory=Person 1')
o.add_argument(r'--ignore-certificate-errors')

website2="https://netmon/Orion/Report.aspx?ReportID=1043&ReturnTo=aHR0cHM6Ly9uZXRtb24vb3Jpb24vcmVwb3J0cy92aWV3cmVwb3J0cy5hc3B4&showid=9eb81515721e448e9ab5381ac71c17c0"
website3="https://netmon/Orion/Login.aspx?ReturnUrl=%2fOrion%2fReport.aspx%3fReportID%3d1041%26ReturnTo%3daHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg%3d&ReportID=1041&ReturnTo=aHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg="
website4="https://netmon/Orion/Login.aspx?ReturnUrl=%2fOrion%2fReport.aspx%3fReportID%3d1039%26ReturnTo%3daHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg%3d&ReportID=1039&ReturnTo=aHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg="
website5="https://amd.service-now.com/navpage.do"
path="C:\\VS code\\chromedriver_win32\\chromedriver"
pdfFileObj = open('C:\\Users\\mukulnag\\Downloads\\a.pdf', 'rb') 
s=Service(path)
driver=webdriver.Chrome(service=s,options=o)


driver.get(website2)
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Username"]').send_keys("amd\mukulnag")
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Password"]').send_keys("45Amd!1101")
b=driver.find_element("xpath",'//a[@id="ctl00_BodyContent_LoginButton"]')
b.click()
c=driver.find_element("xpath",'//a[@id="ExportToExcel"]')
c.click()


driver.get(website3)
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Username"]').send_keys("amd\mukulnag")
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Password"]').send_keys("45Amd!1101")
b=driver.find_element("xpath",'//a[@id="ctl00_BodyContent_LoginButton"]')
b.click()
c=driver.find_element("xpath",'//a[@id="ExportToExcel"]')
c.click()

driver.get(website4)
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Username"]').send_keys("amd\mukulnag")
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Password"]').send_keys("45Amd!1101")
b=driver.find_element("xpath",'//a[@id="ctl00_BodyContent_LoginButton"]')
b.click()
c=driver.find_element("xpath",'//a[@id="ExportToExcel"]')
c.click()

driver.implicitly_wait(20)
time.sleep(5)

"""Clear the html tags
    """
CLEANR = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')

def cleanhtml(raw_html):
  cleantext = re.sub(CLEANR, '', raw_html)
  return cleantext


"""Paths
    """
path = "C:\\Users\\mukulnag\\OneDrive - Advanced Micro Devices Inc\\Documents\\important documents\\VPN and Internet Usage.xlsx"
path2="C:\\Users\\mukulnag\\Downloads\\Report_GP-VPN_Max-Users_last_12_Hrs.xlsx"
path3="C:\\Users\\mukulnag\\Downloads\\Report_AMD_VPN_internet_connection_95th_Percentile_Utilization_-_Last_12_hours.xlsx"
path4="C:\\Users\\mukulnag\\Downloads\\Report_AMD_Non-VPN_internet_connection_95th_Percentile_Utilization_-_Last_12_hours.xlsx"
pathx1 = "C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening-XSJ__Internet_ISP_95__Util.csv"
pathz="C:\\Users\\mukulnag\\Downloads\\LiveNX_-_Scheduled_Report_-_Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening_is_available.zip"
pathx = "C:\\Users\\mukulnag\\Downloads\\"
pathx2 = "C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening-XCO_Internet_ISP_95__Util.csv"
pathx3 ="C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening-XIR_Internet_ISP_95__Util.csv"
pathx4 ="C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening-XHD_Internet_ISP_95__Util.csv"
pathx5 ="C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening-XAP_Internet_ISP_95__Util.csv"
pathx6 ="C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening-XBJ_Internet_ISP_95__Util.csv"


"""Opeing zip for xillinx live report"""
with zipfile.ZipFile(pathz, 'r') as zip_ref:
    zip_ref.extractall(pathx)

"""Open all the work books """

wb1= openpyxl.load_workbook(path) 
wb2= openpyxl.load_workbook(path2)
wb3=openpyxl.load_workbook(path3)
wb5=openpyxl.load_workbook(path4)
s1 = wb1.active
s2=wb2['Sheet1']
s3=wb2['Sheet2']
s4=wb3.active
s5=wb5.active

w1= open(pathx1) 
x1=csv.reader(w1)
w2= open(pathx2) 
x2=csv.reader(w2)
w3= open(pathx3) 
x3=csv.reader(w3)
w4= open(pathx4) 
x4=csv.reader(w4)
w5= open(pathx5) 
x5=csv.reader(w5)
w6= open(pathx6) 
x6=csv.reader(w6)

"""vpnno9=list(s4.iter_rows(min_row=5,max_row=25,min_col=1,max_col=1,values_only=True))
city = list(s1.iter_rows(min_row=5,max_row=25,min_col=1,max_col=1,values_only=True))
values = list(s1.iter_rows(min_row=5,max_row=25,min_col=4,max_col=6,values_only=True))
print(s2.cell(4,4).value)
"""

"""Set up for the vnp % for each site  Vpn 95 input and out put also the non vpn isp"""
vpn95i=[]

for i in range(0,15):
    st=cleanhtml(s4.cell(4+i,2).value)
    st=st.replace(" ", "")
    st=st.replace("%", "")
    vpn95i.append(int(st))
print(vpn95i)
vpn95o=[]

for i in range(0,15):
    st=cleanhtml(s4.cell(4+i,3).value)
    st=st.replace(" ", "")
    st=st.replace("%", "")
    vpn95o.append(int(st))
print(vpn95o)

""" __________________________________________________________________
"""

non95i=[]
for i in range(0,16):
    if i==12:
        continue
    st=cleanhtml(s5.cell(4+i,2).value)
    st=st.replace(" ", "")
    st=st.replace("%", "")
    non95i.append(int(st))
print(non95i)
non95o=[]

for i in range(0,16):
    if i==12:
        continue
    st=cleanhtml(s5.cell(4+i,3).value)
    st=st.replace(" ", "")
    st=st.replace("%", "")
    non95o.append(int(st))
print(non95o)

""" Append the active users for amd and xilinxs sites 
__________________________________________________________________
"""
vpnval=[] 
for k in range(0,24):
    
    vpnval.append(s2.cell(k+4,4).value)

vpnval.append(s3.cell(9,4).value)
vpnval.append(s3.cell(7,4).value)
vpnval.append(s3.cell(8,4).value)
vpnval.append(s3.cell(4,4).value)
vpnval.append(s3.cell(5,4).value)
vpnval.append(s3.cell(6,4).value)

print(vpnval)




h=0
for i in range(5,26):
    for j in range(0,3):
        
        if s1.cell(i,4+j).value=='NA':
            continue
        s1.cell(i,4+j).value=vpnval[h]
        h+=1
h=0

""" For the Xilllinx site liveaction part of report  """
for x in x1:
    if x[1]=="xsj-b101inr01.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i.append(int(x[12]))
        else :
                vpn95o.append(int(x[12]))
                
    elif x[1]=="xsj-d301inr01.xilinx.com": 
        if x[5]=="Inbound":
                non95i.append(int(x[12]))
                
        else :
                non95o.append(int(x[12]))
for x in x2:
    if x[1]=="xco-d110inr02.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i.append(int(x[12]))
        else :
                vpn95o.append(int(x[12]))
                
    elif x[1]=="xco-d110inr01.xilinx.com": 
        if x[5]=="Inbound":
                non95i.append(int(x[12]))
                
        else :
                non95o.append(int(x[12]))
for x in x3:
    if x[1]=="xir-b100inr02.xilinx" :
        if x[5]=="Inbound":
                vpn95i.append(int(x[12]))
        else :
                vpn95o.append(int(x[12]))
                
    elif x[1]=="xir-b100inr01.xilinx": 
        if x[5]=="Inbound":
                non95i.append(int(x[12]))
                
        else :
                non95o.append(int(x[12]))
for x in x4:
    if x[1]=="xhd-d102inr02.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i.append(int(x[12]))
        else :
                vpn95o.append(int(x[12]))
                
    elif x[1]=="xhd-d102inr01.xilinx.com": 
        if x[5]=="Inbound":
                non95i.append(int(x[12]))
                
        else :
                non95o.append(int(x[12]))
for x in x5:
    if x[1]=="xap-d101inr02.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i.append(int(x[12]))
        else :
                vpn95o.append(int(x[12]))
                
    elif x[1]=="xap-d101inr01.xilinx.com": 
        if x[5]=="Inbound":
                non95i.append(int(x[12]))
                
        else :
                non95o.append(int(x[12]))
for x in x6:
    if x[1]=="xbj-b210inr01.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i.append(int(x[12]))
        else :
                vpn95o.append(int(x[12]))
                
    elif x[1]=="xbj-b210inr02.xilinx.com": 
        if x[5]=="Inbound":
                non95i.append(int(x[12]))
                
        else :
                non95o.append(int(x[12]))


""" Append VPN %  lists 
__________________________________________________________________
"""
for i in range(0,21):
    s1.cell(i+5,7).value=vpn95i[h]/100
    s1.cell(i+5,8).value=vpn95o[h]/100
    s1.cell(i+5,9).value=non95i[h]/100
    s1.cell(i+5,10).value=non95o[h]/100
    h+=1

"""For singpore site isp 3"""
st=cleanhtml(s5.cell(16,2).value)
st=st.replace(" ", "")
st=st.replace("%", "")
s1.cell(16,11).value=int(st)/100
st=cleanhtml(s5.cell(16,3).value)
st=st.replace(" ", "")
st=st.replace("%", "")
s1.cell(16,12).value=int(st)/100
print(s1.cell(2,4).value)


"""datetime for report
    """
now = datetime.now()
s1.cell(2,4).value=now.strftime("%m/%d/%y")+"  8:00:00 AM CST "
s1.cell(40,1).value="<Network-Ops@amd.com>,  <Katrin.Schulenburg@amd.com>"
s1.cell(42,1).value=now.strftime("%m/%d")+" 8:00 AM CT Daily Report - Internet and VPN Utilization Monitoring"
s1.cell(43,1).value=" Hi Katrin,\n\nPlease find the VPN utilization report below."
s1.cell(41,1).value=" <DL_InfoSec@amd.com>,  <dl.gis-networking@amd.com>,  <matthews.daniel@amd.com>, <christine.teh@amd.com>,  <suman.chatterjee@amd.com>,  <Ronnie.Fong@amd.com>, <sean.compton@amd.com>,  <mark.hodges@amd.com>, <Michael.Orlando@amd.com>"

print(vpn95i,vpn95o,non95i,non95o) 

driver.quit()



pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 
pageObj = pdfReader.getPage(0) 
pageObj1 = pdfReader.getPage(1) 
ma="(Atlanta|Austin|Bangalore|Beijing|Boxborough|Cyberjaya|Hyderabad|Markham|Munich|Orlando|Santa Clara|Shanghai|Singapore|Taipei|Tokyo|San Jose|Dublin|Longmont|Total).*$"
str=pageObj.extractText()
str+="\n"+pageObj1.extractText()
k=[]
k=str.split("\n")

i=0
for j in range(0,18):        
    s1.cell(5+j,14).value=0
for x in k:
    
    if re.search(ma,x):
        m=re.search(ma,x)
        s=m.group()
        d=re.search("[0-9][0-9]?[0-9]?",s)
        s=s.replace(" ","")
        s=re.sub("\d","",s)
        if d==None:
            continue
        if s=="Total":
            s1.cell(26,15).value=d.group()+" open 'vpn' tickets including ODC site"
            continue
        
        d=int(d.group())
        if s=="Boxborough":
            s1.cell(9,14).value=d
            continue
        if s=="Dublin":
            s1.cell(22,14).value=d
            continue
        if s=="SanJose":
            s1.cell(20,14).value=d
            continue
        if s=="SantaClara":
            s1.cell(15,14).value=d
            continue
        if s=="Longmont":
            s1.cell(21,14).value=d
            continue
        
        for j in range(0,18):
            if s1.cell(5+j,1).value==s:
                s1.cell(5+j,14).value=d
                break


wb1.save("C:\\Users\\mukulnag\\Downloads\\final.xlsx")
wb1.close()
wb2.close()
wb3.close()
wb5.close()
w1.close()
w2.close()
w3.close()
w4.close()
w5.close()
w6.close()
pdfFileObj.close()

os.remove(path2)
os.remove(path3)
os.remove(path4)
os.remove(pathx1)
os.remove(pathx2)
os.remove(pathx3)
os.remove(pathx4)
os.remove(pathx5)
os.remove(pathx6)
os.remove("C:\\Users\\mukulnag\\Downloads\\a.pdf")

os.remove("C:\\Users\\mukulnag\\Downloads\\LiveNX_-_Scheduled_Report_-_Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening_is_available.zip")
os.remove("C:\\Users\\mukulnag\\Downloads\\Legacy_Xilinx_Core_Sites_Internet_Traffic_Stats_-Evening.pdf")
