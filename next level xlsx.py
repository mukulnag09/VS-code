from multiprocessing.connection import wait
from sys import builtin_module_names
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import PyPDF2 
import re
import openpyxl

o = webdriver.ChromeOptions()
o.add_argument(r"--user-data-dir=C:\\Users\\mukulnag\\AppData\\Local\\Google\\Chrome\\User Data\\Default") #e.g. C:\Users\You\AppData\Local\Google\Chrome\User Data
o.add_argument(r'--profile-directory=Person 1')
o.add_argument(r'--ignore-certificate-errors')
pdfFileObj = open('C:\\Users\\mukulnag\\Downloads\\a.pdf', 'rb') 


website5="https://amd.service-now.com/sys_report_template.do?jvar_report_id=bfdbcbdf1b1a81103aefdbd7b04bcbca"
path="C:\\VS code\\chromedriver_win32\\chromedriver"
path2 = "C:\\Users\\mukulnag\\Downloads\\VPN and Internet Usage.xlsx"

"""s=Service(path)
driver=webdriver.Chrome(service=s,options=o)

driver.get(website5)

time.sleep(15)


k=driver.find_element("xpath",'//button[@id="sharing-button"]')
k.click()
time.sleep(5)
k=driver.find_element("xpath",'//a[@id="export-to-pdf-button"]')
k.click()
time.sleep(15)"""



"""driver.quit()"""
wb1= openpyxl.load_workbook(path2)
s1 = wb1.active


pdfReader = PyPDF2.PdfFileReader(pdfFileObj) 
pageObj = pdfReader.getPage(0) 
pageObj1 = pdfReader.getPage(1) 
ma="(Atlanta|Austin|Bangalore|Beijing|Boxborough|Cyberjaya|Hyderabad|Markham|Munich|Orlando|Santa Clara|Shanghai|Singapore|Taipei|Tokyo|San Jose|Dublin|Longmont|Total).*$"
str=pageObj.extractText()
str+="\n"+pageObj1.extractText()
k=[]
k=str.split("\n")

i=0
for j in range(0,18):        
    s1.cell(5+j,14).value=0
for x in k:
    
    if re.search(ma,x):
        m=re.search(ma,x)
        s=m.group()
        d=re.search("[0-9][0-9]?[0-9]?",s)
        s=s.replace(" ","")
        s=re.sub("\d","",s)
        if d==None:
            continue
        if s=="Total":
            s1.cell(26,15).value=d.group()+" open 'vpn' tickets including ODC site"
            continue
        print(s,d.group())
        d=int(d.group())
        if s=="Boxborough":
            s1.cell(9,14).value=d
            continue
        if s=="Dublin":
            s1.cell(22,14).value=d
            continue
        if s=="SanJose":
            s1.cell(20,14).value=d
            continue
        if s=="SantaClara":
            s1.cell(15,14).value=d
            continue
        if s=="Longmont":
            s1.cell(21,14).value=d
            continue
        
        for j in range(0,18):
            if s1.cell(5+j,1).value==s:
                s1.cell(5+j,14).value=d
                break

wb1.save("C:\\Users\\mukulnag\\Downloads\\check.xlsx")
                



#print(k)
