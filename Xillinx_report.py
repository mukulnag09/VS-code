from pickle import APPEND
from sys import builtin_module_names
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import openpyxl
import os
import csv
import zipfile
from datetime import datetime

p=0
"""Clear the html tags_____________________________________________________________________________________________________
    """
CLEANR = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')

def cleanhtml(raw_html):
  cleantext = re.sub(CLEANR, '', raw_html)
  return cleantext

o = webdriver.ChromeOptions()
o.add_argument(r"--user-data-dir=C:\\Users\\mukulnag\\AppData\\Local\\Google\\Chrome\\User Data\\Default") #e.g. C:\Users\You\AppData\Local\Google\Chrome\User Data
o.add_argument(r'--profile-directory=Person 1')
o.add_argument(r'--ignore-certificate-errors')


"""Paths_______________________________________________________________________________________________________"""

website6="https://atlsolmpp01.amd.com/ui/perfstack/PSTK-E3AA90C1D53CCAAADC65C1BD51FE5CC046C3048D"
website3="https://netmon/Orion/Login.aspx?ReturnUrl=%2fOrion%2fReport.aspx%3fReportID%3d1041%26ReturnTo%3daHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg%3d&ReportID=1041&ReturnTo=aHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg="
website4="https://netmon/Orion/Login.aspx?ReturnUrl=%2fOrion%2fReport.aspx%3fReportID%3d1039%26ReturnTo%3daHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg%3d&ReportID=1039&ReturnTo=aHR0cHM6Ly9uZXRtb24vT3Jpb24vUmVwb3J0cy9EZWZhdWx0LmFzcHg="

path="C:\\VS code\\chromedriver_win32\\chromedriver"
path1="C:\\Users\mukulnag\\OneDrive - Advanced Micro Devices Inc\\Documents\\important documents\\Xlinx template.xlsx"
path3="C:\\Users\\mukulnag\\Downloads\\Report_AMD_VPN_internet_connection_95th_Percentile_Utilization_-_Last_12_hours.xlsx"
path4="C:\\Users\\mukulnag\\Downloads\\Report_AMD_Non-VPN_internet_connection_95th_Percentile_Utilization_-_Last_12_hours.xlsx"
path_part2="C:\\Users\\mukulnag\\Downloads\\AMD-XLNX_B2B_95__Daily-Interface_Burstable_Rate.csv"
path_part3="C:\\Users\mukulnag\\OneDrive - Advanced Micro Devices Inc\\Documents\\important documents\\Part 3-4.xlsx"

s=Service(path)
driver=webdriver.Chrome(service=s,options=o)

wbp=openpyxl.load_workbook(path_part3)
sp3=wbp.active

"""Get the reports _________________________________________________________________________________________"""


if(p==1):
    driver.get(website3)
    driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Username"]').send_keys("amd\mukulnag")
    driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Password"]').send_keys("45Amd!1101")
    b=driver.find_element("xpath",'//a[@id="ctl00_BodyContent_LoginButton"]')
    b.click()
    c=driver.find_element("xpath",'//a[@id="ExportToExcel"]')
    c.click()

    driver.get(website4)
    driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Username"]').send_keys("amd\mukulnag")
    driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Password"]').send_keys("45Amd!1101")
    b=driver.find_element("xpath",'//a[@id="ctl00_BodyContent_LoginButton"]')
    b.click()
    c=driver.find_element("xpath",'//a[@id="ExportToExcel"]')
    c.click()



"""________________________________________________________________________________________________________"""

"""Get Part 1 done ....__________________________________________________________________________-
    """

wb1= openpyxl.load_workbook(path1)
s1 = wb1.active

driver.get(website6)
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Username"]').send_keys("amd\mukulnag")
driver.find_element("xpath",'//input[@name="ctl00$BodyContent$Password"]').send_keys("45Amd!1101")
b=driver.find_element("xpath",'//a[@id="ctl00_BodyContent_LoginButton"]')
b.click()
element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located(("xpath", '//div[@class="ps-color-icon ps-color-1"]'))
    )
blue=driver.find_elements("xpath",'//div[@class="ps-color-icon ps-color-1"]')
data1=[]
gbandmb1=[]
iterate=0
for l in blue:
    print(l.text)
    if iterate%2==0 or iterate==0:
        data1.append(int(re.sub("[^0-9]", "",l.text)))
        gbandmb1.append(re.sub("[^a-zA-Z]+", "", l.text))
    iterate+=1
green=driver.find_elements("xpath",'//div[@class="ps-color-icon ps-color-2"]')
data2=[]
gbandmb2=[]
iterate=0
for l in green:
    print(l.text)
    if iterate%2==0 or iterate==0:
        data2.append(int(re.sub("[^0-9]", "",l.text)))
        gbandmb2.append(re.sub("[^a-zA-Z]+", "", l.text))
    iterate+=1
print(data1,data2)


k=0
for x in range(5,10):
    
    if gbandmb1[k]=="G":

        s1.cell(x,3).value=round(data1[k]/1000,2)/10
        print(round(data1[k]/1000,2)/10)
    else:
        s1.cell(x,3).value=round(data1[k]/1000,2)/100
    k+=1
k=0
for x in range(5,10):
    if gbandmb2[k]=="G":
        s1.cell(x,4).value=round(data2[k]/1000,2)/10
        print(round(data2[k]/1000,2)/10)
    else:
        s1.cell(x,4).value=round(data2[k]/1000,2)/100
    k+=1







"""Get Part 2 done ....__________________________________________________________________________-
"""



wb3=openpyxl.load_workbook(path3)
wb5=openpyxl.load_workbook(path4)



try:
    w1= open(path_part2) 
    x1=csv.reader(w1)
except (FileNotFoundError, IOError):
    w1= open("C:\\Users\\mukulnag\\Downloads\\AMD-XLNX_B2B_95__Daily-Interface_Burstable_Rate (1).csv") 
    x1=csv.reader(w1)




s4=wb3.active
s5=wb5.active



vpn95i=[0,0,0,0,0,0,0]
vpn95o=[0,0,0,0,0,0,0]

st=cleanhtml(s4.cell(4,2).value)
st=st.replace(" ", "")
st=st.replace("%", "")
vpn95i[0]=(int(st))
st=cleanhtml(s4.cell(16,2).value)
st=st.replace(" ", "")
st=st.replace("%", "")
vpn95i[1]=(int(st))


st=cleanhtml(s4.cell(4,3).value)
st=st.replace(" ", "")
st=st.replace("%", "")
vpn95o[0]=(int(st))
st=cleanhtml(s4.cell(16,3).value)
st=st.replace(" ", "")
st=st.replace("%", "")
vpn95o[1]=(int(st))


non95i=[0,0,0,0,0,0,0]
non95o=[0,0,0,0,0,0,0]

st=cleanhtml(s5.cell(4,2).value)
st=st.replace(" ", "")
st=st.replace("%", "")
non95i[0]=(int(st))
st=cleanhtml(s5.cell(17,2).value)
st=st.replace(" ", "")
st=st.replace("%", "")
non95i[1]=(int(st))


st=cleanhtml(s5.cell(4,3).value)
st=st.replace(" ", "")
st=st.replace("%", "")
non95o[0]=(int(st))
st=cleanhtml(s5.cell(17,3).value)
st=st.replace(" ", "")
st=st.replace("%", "")
non95o[1]=(int(st))
    
for x in x1:
    
    if x[1]=="xsj-b101inr01.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i[2]=int(x[12])
        else :
                vpn95o[2]=int(x[12])
                
    elif x[1]=="xsj-d301inr01.xilinx.com": 
        if x[5]=="Inbound":
                non95i[2]=int(x[12])
                
        else :
                non95o[2]=int(x[12])

   
    if x[1]=="xco-d110inr01.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i[4]=int(x[12])
        else :
                vpn95o[4]=int(x[12])
                
    elif x[1]=="xco-d110inr02.xilinx.com": 
        if x[5]=="Inbound":
                non95i[4]=int(x[12])
                
        else :
                non95o[4]=int(x[12])

    if x[1]=="xir-b100inr01.xilinx" :
        if x[5]=="Inbound":
                vpn95i[5]=int(x[12])
        else :
                vpn95o[5]=int(x[12])
                
    elif x[1]=="xir-b100inr02.xilinx": 
        if x[5]=="Inbound":
                non95i[5]=int(x[12])
                
        else :
                non95o[5]=int(x[12])

    if x[1]=="xhd-d102inr01.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i[6]=int(x[12])
        else :
                vpn95o[6]=int(x[12])
                
    elif x[1]=="xhd-d102inr02.xilinx.com": 
        if x[5]=="Inbound":
                non95i[6]=int(x[12])
                
        else :
                non95o[6]=int(x[12])

    if x[1]=="xap-d101inr02.xilinx.com" :
        if x[5]=="Inbound":
                vpn95i[3]=int(x[12])
        else :
                vpn95o[3]=int(x[12])
                
    elif x[1]=="xap-d101inr01.xilinx.com": 
        if x[5]=="Inbound":
                non95i[3]=int(x[12])
                
        else :
                non95o[3]=int(x[12])

print(vpn95i,vpn95o,non95i,non95o)
for x in range(0,7):
    s1.cell(x+19,3).value=vpn95i[x]/100
    s1.cell(x+19,4).value=vpn95o[x]/100
    s1.cell(x+19,5).value=non95i[x]/100
    s1.cell(x+19,6).value=non95o[x]/100




wb3.close()
wb5.close()
os.remove(path3)
os.remove(path4)



"""part 3 - 4______________________________________________________________________________________________________________"""

laten=[]
loss=[]
laten.append(sp3.cell(5,10).value)
if (sp3.cell(5,13).value*100)>0.99999999:
        loss.append(sp3.cell(5,13).value*100)
else:
        loss.append(0)
for x in range(1,35):
    laten.append(sp3.cell(x*4+5,10).value)
    if (sp3.cell(x*4+5,13).value*100)>0.99999999:
        loss.append(sp3.cell(x*4+5,13).value)
    else:
        loss.append(0)

print(laten)
print(loss)


for x in range(0,7):
    for k in range(0,5):
        s1.cell(k+32,x+2).value=round(laten[x*5+k],0)
        s1.cell(k+44,x+2).value=loss[x*5+k]


wb1.save("C:\\Users\\mukulnag\\Downloads\\xlinx_Report.xlsx")

driver.quit()
